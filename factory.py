from openpyxl import load_workbook
from openpyxl import Workbook
from random import random

LOG = []


class Factory:
    """Factory get bombs and creates minions workflow on conveyors
    """

    def __init__(self, inspect, number_minions, filename):
        """ Factory is set up using parameters:
        1. "inspect" - number of minions, which should test the bomb
        2. "number_minions" - number of minions at the factory
        3. "filename" - file where stored the bombs
        """
        self.inspections = inspect
        self.number_minions = number_minions
        self.bombs = []
        self.minions = []
        self._create_minions(number_minions)
        self._get_bombs(filename)

    def _create_minions(self, number_minions):
        self.minions = [Minion(id) for id in
                        range(1, number_minions + 1)]

    def _get_bombs(self, filename):
        """Get the bombs from the file"""
        wb = load_workbook(filename=filename)
        sheet_range = wb['Bombs']
        row = 2
        while True:
            id = sheet_range["A{0}".format(row)].value
            is_broken = sheet_range['B{0}'.format(row)].value
            if not id:
                break
            self.bombs.append(Bomb(id, is_broken))
            row += 1

    def run_conveyor(self):
        """ Starts circular conveyor """
        count = 0
        while count <= len(self.bombs) / self.number_minions:
            tape = self.bombs[
                   self.number_minions * count:self.number_minions * (
                       count + 1)]
            for check in xrange(self.inspections):
                for i in xrange(len(tape)):
                    self.minions[i - check].check_bomb(tape[i])
            count += 1


class ReportsManager:
    def __init__(self, filename, factory):
        self.filename = filename
        self.factory = factory
        self.motivation_report = {id: [0, 0, 0] for id in
                                  range(1, factory.number_minions + 1)}
        self.qa_report = {}
        self.percentage_correct = 0

    def generate_reports(self):
        """ Generate qa report and motivation report"""
        correctly = 0
        for bomb in self.factory.bombs:
            # generate qa report
            answers = bomb.minions_stickers.values()
            yes = answers.count('yes')
            quorum = yes and float(
                yes) / self.factory.inspections > 0.5 and 'yes' or 'no'
            # The code below count number of bomb, which broken but minions
            # quorum answered 'yes'
            # if quorum == 'yes' and bomb.is_broken:
            #     count += 1
            if quorum == 'yes':
                self.qa_report[bomb.id] = 0
                if not bomb.is_broken:
                    correctly += 1
            else:
                self.qa_report[bomb.id] = 1
                if bomb.is_broken:
                    correctly += 1
            # generate motivation report
            for id_minion, answer in bomb.minions_stickers.iteritems():
                if answer is None:
                    self.motivation_report[id_minion][2] += 1
                    continue
                if answer == quorum:
                    self.motivation_report[id_minion][0] += 1
                else:
                    self.motivation_report[id_minion][1] += 1

        self.percentage_correct = float(correctly) / len(
            self.factory.bombs) * 100

    def write_xlsx(self):
        """ Writes a reports and log in xlsx file"""
        wb = Workbook()

        # write Log
        ws1 = wb.active
        ws1.title = 'Log'
        column_name = ['minion_id', 'bomb_id', 'answer']

        for row in range(1, len(LOG) + 2):
            for col in range(1, 4):
                if row == 1:
                    ws1.cell(column=col, row=row, value=column_name[col - 1])
                    continue
                ws1.cell(column=col, row=row, value=LOG[row - 2][col - 1])

        # write motivation report
        ws2 = wb.create_sheet(title='Motivation Report')
        column_name = ['minion_id', 'banans_cnt', 'flogging_cnt',
                       'bombs_skipped']
        for row in range(0, len(self.motivation_report) + 1):
            for col in range(1, len(column_name) + 1):
                if row == 0:
                    ws2.cell(column=col, row=row + 1,
                             value=column_name[col - 1])
                    continue
                if col == 1:
                    ws2.cell(column=col, row=row + 1, value=row)
                    continue
                ws2.cell(column=col, row=row + 1,
                         value=self.motivation_report[row][col - 2])
        # write qa report
        ws3 = wb.create_sheet(title='QA Report')
        column_name = ['bomb_id', 'is_broken']
        for row in range(0, len(self.qa_report) + 1):
            for col in range(1, len(column_name) + 1):
                if row == 0:
                    ws3.cell(column=col, row=row + 1,
                             value=column_name[col - 1])
                    continue
                if col == 1:
                    ws3.cell(column=col, row=row + 1, value=row)
                    continue
                ws3.cell(column=col, row=row + 1,
                         value=self.qa_report[row])
        ws3.cell(column=3, row=1, value="% of correct")
        ws3.cell(column=3, row=2, value=self.percentage_correct)

        # save file
        wb.save(filename=self.filename)


class Bomb:
    """Model describe bomb with minions stickers."""

    def __init__(self, id, is_broken):
        self.id = id
        self.is_broken = is_broken
        self.minions_stickers = {}


class Minion:
    """Model describe minions.
    Minions can check bomb (method check_bomb).
    WEIGHTS show probability ANSWERS
    """

    ANSWERS = [True, False, None]
    WEIGHTS = [6, 3, 1]

    def __init__(self, id):
        self.id = id

    def check_bomb(self, bomb):
        decision = generate_decision(self.ANSWERS, self.WEIGHTS)
        if decision is None:
            bomb.minions_stickers[self.id] = None
            LOG.append((self.id, bomb.id, None))
            return
        if decision:
            answer = bomb.is_broken and 'no' or 'yes'
            bomb.minions_stickers[self.id] = answer
            LOG.append((self.id, bomb.id, answer))
        else:
            answer = bomb.is_broken and 'yes' or 'no'
            bomb.minions_stickers[self.id] = answer
            LOG.append((self.id, bomb.id, answer))


def generate_decision(answers, weights):
    """Generate minions decision, depends on  possible answers and
    probability this answers
    """
    total_weight = float(sum(weights))
    rel_weight = [w / total_weight for w in weights]

    # Probability for each element
    probs = [sum(rel_weight[:i + 1]) for i in range(len(rel_weight))]

    for (i, element) in enumerate(answers):
        if random() <= probs[i]:
            break
    return element


if __name__ == '__main__':
    factory = Factory(10, 102, 'bombs.xlsx')
    factory.run_conveyor()
    reports_manager = ReportsManager('report.xlsx', factory)
    reports_manager.generate_reports()
    reports_manager.write_xlsx()
