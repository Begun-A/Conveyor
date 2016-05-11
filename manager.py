from openpyxl import load_workbook, Workbook


class WorkbookManager:
    def __init__(self, filename, sheet):
        self.wb = load_workbook(filename)
        self.bombs_sheet = sheet
        self.wb_output = Workbook(filename)
        self.log_sheet = self.wb.create_sheet(title='Log')
        self.row_cursor = 1

    def get_bombs_data(self):
        bombs_sheet = self.wb.get_sheet_by_name(self.bombs_sheet)
        for row in bombs_sheet.rows[1:]:
            yield [cell.value for cell in row]

    def write_xslx(self, sheet_name, column_names, data):
        title_row = 1
        ws = self.wb.create_sheet(title=sheet_name)
        for col, col_name in enumerate(column_names, 1):
            ws.cell(column=col, row=title_row, value=col_name)

        for row, record in enumerate(data, title_row + 1):
            for col, value in enumerate(record, 1):
                ws.cell(column=col, row=row, value=value)

    def minions_log(self, record):
        """
        Write minion bomb check data to the file.
        """
        column_names = ['minion_id', 'bomb_id', 'answer']
        if self.row_cursor == 1:
            for col, col_name in enumerate(column_names, 1):
                self.log_sheet.cell(column=col, row=self.row_cursor, value=col_name)
            self.row_cursor += 1
        for col, value in enumerate(record, 1):
            self.log_sheet.cell(row=self.row_cursor, column=col, value=value)
        self.row_cursor += 1