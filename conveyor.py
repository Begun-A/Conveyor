from random import random
from openpyxl import load_workbook, Workbook
import argparse

class Bomb:
    """Model describe bomb with minions stickers."""

    def __init__(self, id, is_broken):
        self.id = id
        self.is_broken = is_broken
        self.minions_stickers = {}


class Conveyor:
    """Conveyor get bombs and creates minions workflow on conveyors"""

    def __init__(self, number_minions, minions_qualific, inspect, filename,
                 sheet):
        """Initialize a Conveyor instance.

        Parameters:
        number_minions -- number of minions at the conveyor
        minions_qualific -- the probability of a correct answer minions.
                            Is defined as a tuple = (right, wrong, skip)
        inspect -- number of minions, which should test the bomb
        filename -- xlsx file with bombs
        sheet -- sheet, where stored the bombs.
        """
        self.number_minions = number_minions
        self.weights = minions_qualific
        self.inspect = inspect
        self.filename = filename
        self.sheet = sheet
        self.wb_manager = WorkbookManager(filename, sheet)
        self.bombs = [Bomb(*record) for record in
                      self.wb_manager.get_bombs_data()]
        self.minions = [Minion(id, minions_qualific, self.wb_manager) for id in
                        range(1, number_minions+1)]
        self.motivation_report = []
        self.qa_report = []
        self.percentage_correct = 0

    def run_conveyor(self):
        """ Starts circular conveyor """
        count = 0
        while count <= len(self.bombs) / self.number_minions:
            from_box = self.number_minions * count
            to_box = self.number_minions * (count + 1)
            tape = self.bombs[from_box:to_box]
            for turn in xrange(self.inspect):
                for pos in xrange(len(tape)):
                    self.minions[pos - turn].check_bomb(tape[pos])
            count += 1

    def generate_reports_and_salary(self):
        """ Generate qa report and motivation report"""
        correctly = 0.
        for bomb in self.bombs:
            # generate qa report
            answers = bomb.minions_stickers.values()
            yes = float(answers.count('yes'))
            quorum = 'yes' if yes / self.inspect > 0.5 else 'no'
            if quorum == 'yes':
                self.qa_report.append((bomb.id, 0))
                correctly += 1 if not bomb.is_broken else 0
            else:
                self.qa_report.append((bomb.id, 1))
                correctly += 1 if bomb.is_broken else 0
            # generate salary
            for minion, answer in bomb.minions_stickers.iteritems():
                if answer is None:
                    minion.skips += 1
                    continue
                elif answer == quorum:
                    minion.bananas += 1
                else:
                    minion.floggings += 1
        # generate motivation reports
        for minion in self.minions:
            self.motivation_report.append((minion.id, minion.bananas,
                                           minion.floggings, minion.skips))
        self.percentage_correct = correctly / len(self.bombs) * 100

    def write_reports(self):
        """ Writes a reports and log in xlsx file"""

        # write motivation report
        column_names = ['minion_id', 'bananas_cnt', 'flogging_cnt',
                        'bombs_skipped']
        data = self.motivation_report
        self.wb_manager.write_xslx('Motivation Report', column_names, data)

        # write qa report
        column_names = ['bomb_id', 'is_broken']
        self.wb_manager.write_xslx('QA Report', column_names, self.qa_report)

        # write percentage correct
        column_names = ['% correct']
        data = [(self.percentage_correct,)]
        self.wb_manager.write_xslx('% correct', column_names, data)

        # save in output file
        filename = self.filename
        ind = filename.find('.')
        output_file = filename[:ind] + '_output' + filename[ind:]
        self.wb_manager.wb.save(filename=output_file)


class Minion:
    """
    Model describe minions.
    Minions can check bomb (method check_bomb).
    """

    ANSWERS = (True, False, None)

    def __init__(self, id, weights, wb_manager):
        """Initialize a Conveyor instance.

        Parameters:
        "weights" - show probability ANSWERS
        """
        self.id = id
        self.weights = weights
        self.wb_manager = wb_manager
        self.bananas = 0
        self.floggings = 0
        self.skips = 0

    def check_bomb(self, bomb):
        decision = self.generate_decision(Minion.ANSWERS, self.weights)
        if decision is None:
            answer = None
        elif decision:
            answer = 'no' if bomb.is_broken else 'yes'
        else:
            answer = 'yes' if bomb.is_broken else 'no'
        bomb.minions_stickers[self] = answer
        self.wb_manager.minions_log((self.id, bomb.id, answer))

    @staticmethod
    def generate_decision(answers, weights):
        """
        Generate minions decision, depends on  possible answers and
        probability this answers
        """
        total_weight = float(sum(weights))
        rel_weight = [w / total_weight for w in weights]

        # Probability for each element
        probs = [sum(rel_weight[:i + 1]) for i in range(len(rel_weight))]

        for (i, element) in enumerate(answers):
            if random() <= probs[i]:
                break
        return element


class WorkbookManager:
    def __init__(self, filename, sheet):
        self.wb = load_workbook(filename)
        self.bombs_sheet = sheet
        self.wb_output = Workbook(filename)
        self.log_sheet = self.wb.create_sheet(title='Log')
        self.row_cursor = 1

    def get_bombs_data(self):
        bombs_sheet = self.wb.get_sheet_by_name(self.bombs_sheet)
        for row in bombs_sheet.rows[1:]:
            yield [cell.value for cell in row]

    def write_xslx(self, sheet_name, column_names, data):
        title_row = 1
        ws = self.wb.create_sheet(title=sheet_name)
        for col, col_name in enumerate(column_names, 1):
            ws.cell(column=col, row=title_row, value=col_name)

        for row, record in enumerate(data, title_row + 1):
            for col, value in enumerate(record, 1):
                ws.cell(column=col, row=row, value=value)

    def minions_log(self, record):
        """
        Write minion bomb check data to the file.
        """
        column_names = ['minion_id', 'bomb_id', 'answer']
        if self.row_cursor == 1:
            for col, col_name in enumerate(column_names, 1):
                self.log_sheet.cell(column=col, row=self.row_cursor, value=col_name)
            self.row_cursor += 1
        for col, value in enumerate(record, 1):
            self.log_sheet.cell(row=self.row_cursor, column=col, value=value)
        self.row_cursor += 1


if __name__ == '__main__':
    # parsing arguments
    import sys
    parser = argparse.ArgumentParser()
    parser.add_argument('-m', action='store', dest='minions',
                        help='Number of minions on conveyor')
    parser.add_argument('-c', action='store', dest='correct', default=60,
                        help='The probability of a correct answer. Defult=60')
    parser.add_argument('-s', action='store', dest='skip', default=10,
                        help='The probability of a skip')
    parser.add_argument('-i', action='store', dest='inspect',
                        help='Number of inspections each bomb')
    parser.add_argument('-file', action='store', dest='file',
                        default='bombs.xlsx', help='Xlsx file with bombs')
    parser.add_argument('-sheet', action='store', dest='sheet', default='Bombs',
                        help='Shee with bombs in xlsx file')
    args = parser.parse_args()

    correct = float(args.correct) / 10
    skip = float(args.skip) / 10
    if correct + skip > 100:
        sys.exit('The sum of probabilities of  correct and skip must be < 100')
    if not args.minions or not args.inspect:
        sys.exit('Number of minions and inspections is required')
    minions = int(args.minions)
    inspect = int(args.inspect)
    if minions < inspect:
        sys.exit('Number of inspections can\'t be more then minions')
    qualific = (correct, 10 - correct - skip, skip)

    # program execution
    conveyor = Conveyor(
        minions,
        qualific,
        inspect,
        args.file,
        args.sheet
    )
    conveyor.run_conveyor()
    conveyor.generate_reports_and_salary()
    conveyor.write_reports()
